import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from datetime import datetime

# Define the file path (Modify as needed)
excel_file = r"sample_comparison.xlsx"

# Load the Excel file
wb = load_workbook(excel_file)
if "Revised" not in wb.sheetnames or "Actual" not in wb.sheetnames:
    print("Error: 'Revised' or 'Actual' sheet not found in the file.")
    exit()

ws_revised = wb["Revised"]
ws_actual = wb["Actual"]

# Get maximum rows and columns
max_row = max(ws_revised.max_row, ws_actual.max_row)
max_col = max(ws_revised.max_column, ws_actual.max_column)

# Define fill colors
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for match
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for mismatch

# Function to convert cell values into comparable format
def convert_value(value):
    if isinstance(value, datetime):  # Handle date format
        return value.strftime("%Y-%m-%d")
    elif isinstance(value, (int, float)):  # Ensure numbers are properly compared
        return round(value, 3)  # Keep precision up to 3 decimal places
    elif value is None:  # Handle empty values
        return ""
    else:
        return str(value).strip()  # Convert everything else to string for comparison

# Iterate cell by cell
for row in range(2, max_row + 1):  # Start from 2 to skip headers
    for col in range(1, max_col + 1):
        cell_revised = ws_revised.cell(row=row, column=col)
        cell_actual = ws_actual.cell(row=row, column=col)

        value_revised = convert_value(cell_revised.value)
        value_actual = convert_value(cell_actual.value)

        if value_revised == value_actual:
            cell_revised.fill = green_fill
            cell_actual.fill = green_fill
        else:
            cell_revised.fill = red_fill
            cell_actual.fill = red_fill
            # Add comment with Actual & Revised values
            comment_text = f"Actual: {value_actual}, Revised: {value_revised}"
            cell_revised.comment = Comment(comment_text, "Comparison Script")

# Apply auto-filter to headers
for sheet in [ws_revised, ws_actual]:
    sheet.auto_filter.ref = sheet.dimensions  # Apply filter to all headers

# Save the updated Excel file
wb.save(excel_file)
print(f"Comparison complete! File saved: {excel_file}")
