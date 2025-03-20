import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_excel_files(file_path, sheet1, sheet2, key_column):
    print(f"Processing: {file_path}")
    compare_excel_sheets(file_path, sheet1, sheet2, key_column)

def compare_excel_sheets(file_path, sheet1, sheet2, key_column):
    # Load both sheets into pandas DataFrames
    df_before = pd.read_excel(file_path, sheet_name=sheet1, dtype={"LDCNo": str})
    df_revised = pd.read_excel(file_path, sheet_name=sheet2, dtype={"LDCNo": str})
    
    # Sort both DataFrames by the key column before comparison
    df_before = df_before.sort_values(by=[key_column]).reset_index(drop=True)
    df_revised = df_revised.sort_values(by=[key_column]).reset_index(drop=True)
    
    # Ensure both DataFrames have the same structure
    if not df_before.columns.equals(df_revised.columns):
        raise ValueError(f"Column structures in {file_path} do not match!")
    
    # Sum EnergyMouVol values
    energy_mou_vol_sum_before = df_before["EnergyMouVol"].sum()
    energy_mou_vol_sum_revised = df_revised["EnergyMouVol"].sum()
    print(f"Total EnergyMouVol in {file_path} - Before: {energy_mou_vol_sum_before}, Revised: {energy_mou_vol_sum_revised}")
    
    # Load workbook to update sheets
    wb = load_workbook(file_path)
    ws_before = wb[sheet1]
    ws_revised = wb[sheet2]
    
    # Find last row with data and write total sum after one blank row
    last_row_before = ws_before.max_row + 1
    last_row_revised = ws_revised.max_row + 1
    
    ws_before[f"C{last_row_before+1}"] = "Total EnergyMouVol"
    ws_before[f"D{last_row_before+1}"] = energy_mou_vol_sum_before
    ws_revised[f"C{last_row_revised+1}"] = "Total EnergyMouVol"
    ws_revised[f"D{last_row_revised+1}"] = energy_mou_vol_sum_revised
    
    # Compare row by row and track differences
    differences = []
    max_rows = max(len(df_before), len(df_revised))
    
    for index in range(max_rows):
        if index >= len(df_before) or index >= len(df_revised):
            continue  # Skip if one sheet has fewer rows than the other
        
        data_before = df_before.iloc[index]
        data_revised = df_revised.iloc[index]
        
        for col in df_before.columns:
            if data_before[col] != data_revised[col]:
                differences.append({
                    "Row Number": index + 2,  # Excel row number (1-based index, +1 for header)
                    "Column Name": col,
                    "Before Value": data_before[col],
                    "Revised Value": data_revised[col]
                })
    
    # Save differences in a new sheet or overwrite if exists
    if differences:
        diff_df = pd.DataFrame(differences)
        with pd.ExcelWriter(file_path, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer:
            diff_df.to_excel(writer, sheet_name="Differences", index=False)
    
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    
    # Highlight differences in red
    for diff in differences:
        row_num = diff["Row Number"]
        col_letter = ws_before.cell(row=1, column=list(df_before.columns).index(diff["Column Name"]) + 1).column_letter
        
        # Apply red highlight to mismatched rows in both sheets
        for ws in [ws_before, ws_revised]:
            ws[f"{col_letter}{row_num}"].fill = red_fill
    
    # Highlight total EnergyMouVol comparison
    total_row = last_row_before + 1
    total_col_letter = "D"
    if energy_mou_vol_sum_before == energy_mou_vol_sum_revised:
        ws_before[f"{total_col_letter}{total_row}"].fill = green_fill
        ws_revised[f"{total_col_letter}{total_row}"].fill = green_fill
    else:
        ws_before[f"{total_col_letter}{total_row}"].fill = red_fill
        ws_revised[f"{total_col_letter}{total_row}"].fill = red_fill
    
    # Save the changes
    wb.save(file_path)
    print(f"Comparison complete for {file_path}. Differences sheet updated and mismatched rows highlighted.")

# Example usage
file_path = "C:\\test\\sample_excel.xlsx"
process_excel_files(file_path, "Before", "Revised", "CustomerNo")
