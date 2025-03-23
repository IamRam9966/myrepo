import pyodbc
import pandas as pd
import os
import time
import uuid
from openpyxl import load_workbook
from openpyxl.worksheet.filters import AutoFilter

# Azure SQL Server Connection Details
server = "your-server-name.database.windows.net"
database = "your-database-name"
username = input("Enter your Azure SQL username: ")

# Establish Connection using Azure MFA
conn_str = f"DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={server};DATABASE={database};AUTHENTICATION=ActiveDirectoryInteractive;UID={username}"
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

# Define customer details (customer_no, customer_name)
customers = [
    (1, "Herman Miller"),
    (2, "CBS Broadcasting"),
    # Add more customers as needed
]

# Define output folder
output_folder = r"C:\Users\Ram\Documents\Generated"
os.makedirs(output_folder, exist_ok=True)

# Function to execute queries
def get_query_results(query):
    try:
        print(f"Executing query: {query}")  # Debugging
        cursor.execute(query)
        
        if cursor.description is None:
            print(f"Warning: No results found for query: {query}")
            return None
        
        columns = [desc[0] for desc in cursor.description]
        data = cursor.fetchall()
        
        return pd.DataFrame.from_records(data, columns=columns) if data else None
    except Exception as e:
        print(f"Error executing query: {query}: {e}")
        return None

# Function to ensure file is closed before writing
def ensure_file_closed(filepath):
    max_retries = 5
    for attempt in range(max_retries):
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
            return
        except PermissionError:
            print(f"Attempt {attempt + 1}: File {filepath} is in use. Close it and retrying in 3 seconds...")
            time.sleep(3)
    print(f"Failed to delete {filepath}. Ensure it is closed and try again.")

# Loop through customers and fetch data
for customer_no, customer_name in customers:
    unique_id = uuid.uuid4()
    sanitized_no = str(customer_no).replace(" ", "")
    sanitized_name = customer_name.replace(" ", "").replace("/", "")  # Ensure filename is safe
    excel_file = os.path.join(output_folder, f"customer_data_{sanitized_no}_{sanitized_name}_{unique_id}.xlsx")
    
    # Ensure the file is closed before opening in write mode
    ensure_file_closed(excel_file)

    # Replace ? with actual customer_no in queries
    query1 = f"""
        SELECT * 
        FROM RevisedData
        WHERE CustomerNo = {customer_no}
    """
    
    query2 = f"""
        SELECT * 
        FROM BeforeData
        WHERE CustomerNo = {customer_no}
    """
    
    df1 = get_query_results(query1)
    if df1 is None:
        continue
    
    df2 = get_query_results(query2)
    if df2 is None:
        continue
    
    # Write to Excel using openpyxl
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
        sheets_written = 0
        
        for df, sheet_name in [(df1, "Revised"), (df2, "Before")]:
            if not df.empty:
                energy_col = "EnergyMUVol"
                if energy_col in df.columns:
                    df[energy_col] = df[energy_col].astype(float).round(3)
                    sum_value = round(df[energy_col].sum(), 3)
                    blank_row = pd.DataFrame([{col: None for col in df.columns}])
                    sum_row = pd.DataFrame([{col: None for col in df.columns}])
                    sum_row[energy_col] = sum_value  # Keep as float, not string
                    df = pd.concat([df, blank_row, sum_row], ignore_index=True)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                sheets_written += 1
        
        # Ensure at least one sheet is written to avoid Excel error
        if sheets_written == 0:
            pd.DataFrame({"Message": ["No data available for this customer"]}).to_excel(writer, sheet_name="Placeholder", index=False)
    
    # Load workbook and apply filter to header
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions  # Apply filter to all columns in the header
    wb.save(excel_file)
    
    print(f"Excel file generated: {excel_file}")

cursor.close()
conn.close()
